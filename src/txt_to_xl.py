from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import ColorScaleRule
import os

wb = Workbook()

def add_songs(sheet: Worksheet, directory: str, filename: str) -> int:
    mxw = 0

    with open(f"{directory}/{filename}.txt", encoding='UTF-8') as src:
        idx = 2
        for line in src.readlines():
            sheet[f'A{idx}'] = line
            mxw = max(mxw, len(line))
            idx+=1
    
    sheet.column_dimensions['A'].width = mxw+2
    return idx-2

def add_score_table(sheet: Worksheet, n: int, filename: str):
    names = filename.split("-")
    
    # First Column
    line = 2
    for name in names:
        sheet[f'E{line}'] = name
        line += 1
    sheet[f'E{line}'] = 'Ninguém'

    # First Row
    sheet['F1'] = "Nomes"
    sheet['G1'] = "Artistas"
    sheet['H1'] = "Total"
    sheet['I1'] = "Pontuação"

    # General Formulas
    for i in range(2, line):
        letra = sheet[f'E{i}'].value[0]
        sheet[f'F{i}'] = f"=COUNTIF(B2:B{n+1}, \"{letra}\")" #### TODODODODODOD
        sheet[f'G{i}'] = f"=COUNTIF(C2:C{n+1}, \"{letra}\")"
        sheet[f'H{i}'] = f"=F{i}+G{i}"
        sheet[f'I{i}'] = f"=H{i}/SUM(H2:H{line})"

    # Last Formulas
    sheet[f'F{line}'] = f"=COUNTIF(B2:B{n+1}, \"\")"
    sheet[f'G{line}'] = f"=COUNTIF(C2:C{n+1}, \"\")"
    sheet[f'H{line}'] = f"=F{line}+G{line}"
    sheet[f'I{line}'] = f"=H{line}/SUM(H2:H{line})"

def populate_sheet(sheet: Worksheet, directory: str, filename: str):
    sheet['B1'] = "Nome"
    sheet['C1'] = "Artista"

    n = add_songs(sheet, directory, filename)

    add_score_table(sheet, n, filename)

def create_totals(sheet: Worksheet, names: list[str], sheetnames: list[str]):
    sheet.title = "SCORES"

    formula = "=SUMIF('{sheetname}'!E2:E20, SCORES!A{line}, '{sheetname}'!I2:I20)"
    #formula = "=" + " + ".join(["SUMIF('" + sn + "'!E2:E20, SCORES!A{line}, '" + sn + "'!I2:I20)" for sn in sheetnames if sn != "Sheet"])

    first_row: tuple[Cell] = sheet.iter_rows(1, 1, 3, len(sheetnames)+1).__next__()
    for i, cell in enumerate(first_row):
        cell.value = sheetnames[i+1]
    
    first_col: tuple[Cell] = sheet.iter_cols(1, 1, 2, len(names)+1).__next__()
    for i, cell in enumerate(first_col):
        cell.value = names[i]

    for col in sheet.iter_cols(3, len(sheetnames)+1, 2, len(names)+1):
        for cell in col:
            cell.value = formula.format(sheetname=sheet[f'{cell.column_letter}1'].value, line=cell.row)
    
    col: tuple[Cell] = sheet.iter_cols(2, 2, 1, len(names)+1).__next__()
    for cell in col:
        if cell.row == 1:
            cell.value = "TOTAL"
            continue
        last = cell.offset(column=len(sheetnames)-1)
        cell.value = f"=SUM(C{cell.row}:{last.column_letter}{cell.row})"

    color_scale_rule = ColorScaleRule(
        start_type="min", start_color="FF0000",  # Red for minimum
        mid_type="percentile", mid_value=99, mid_color="00FF00",  # Yellow at 50th percentile
        end_type="max", end_color="0000FF"  # Green for maximum
    )

    # Add the rule to the worksheet
    sheet.conditional_formatting.add(f"B2:B{len(names)+1}", color_scale_rule)
        
def create_score_table(directory: str, usernames: list[str]):
    for file in os.listdir(directory):
        if not file.endswith(".txt"):
            continue
        file = file[:-4]
        compressed = "-".join([name[:3] for name in file.split('-')])
        ws: Worksheet = wb.create_sheet(compressed)
        populate_sheet(ws, directory, file)

    create_totals(wb['Sheet'], usernames, wb.sheetnames)


    wb.save(f"{directory}/scores.xlsx")
    wb.close()
    print(f"Excel file '{directory}/scores.xlsx' has been created.")
